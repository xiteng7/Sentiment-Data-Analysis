#!/usr/bin/env python
# coding: utf-8

# In[20]:


import pandas as pd
from pytube import YouTube
import os
os.environ["IMAGEIO_FFMPEG_EXE"] = "/Users/xiteng/Downloads/ffmpeg"
from moviepy.editor import *


# In[56]:


# this directory saves downloaded videos
os.chdir('/Users/xiteng/Desktop/video sentiment analysis materials/Models/Video-Sentiment-Analysis-master/video data')


# In[37]:


# load excel sp500 data
from openpyxl import load_workbook
wb = load_workbook('sp500(with time intervals).xlsx')
ws = wb.active


# In[43]:


# loop through the column M in excel "sp500(with time intervals).xlsx"
# download the videos and store them in current directory
# set min_row = 3, max_row = 466 to download the full list of videos
lst = []
for row in ws.iter_cols(min_row=3, max_row=4, min_col= 1, max_col=1):
    for cell in row:
        if cell.value == None:
            continue
#         print(cell.value)
        lst.append(cell.value+'.MP4')
print(lst)

for row in ws.iter_cols(min_row=3, max_row=4, min_col= 13, max_col=13):
    for cell, name, in zip(row, lst):
        if cell.value == None:
            continue
        print(cell.value)
        yt = YouTube(cell.value)
        video = yt.streams.get_highest_resolution().download(filename = name)


# In[ ]:


# codes of following blocks used to extract video clips with pre-set time intervals


# In[57]:


# cretae a dataframe of time intervals
df_time_intervals = pd.read_excel('sp500(with time intervals).xlsx', 
                   usecols='O:P', skiprows = 1, nrows = 2, names = ['time_start', 'time_end'])
print(df_time_intervals)

#create lists of start time and end time
time_starts = []
time_ends = []
for i in df_time_intervals.index:
    time_starts.append(df_time_intervals['time_start'][i])
    time_ends.append(df_time_intervals['time_end'][i])
print(time_starts)
print(time_ends)

#create a dictionary with keys=stock ticks, values=start time, end time
#this dictionary is used as a map to loop through each video and extract segments of videos
video_dict = {}
for tick, start, end in zip(lst, time_starts, time_ends):
    video_dict[tick] = [start, end]
print(video_dict)


# In[53]:


#transform downloaded vidoes into clips 
#get videos from 'video_path'
#write transferred clips into 'clips_path'
clips_path = '/Users/xiteng/Desktop/video sentiment analysis materials/Models/Video-Sentiment-Analysis-master/clips'
video_path = '/Users/xiteng/Desktop/video sentiment analysis materials/Models/Video-Sentiment-Analysis-master/video data'
for tick in video_dict:
    os.chdir(video_path)
    clips = VideoFileClip(tick).subclip(video_dict[tick][0],video_dict[tick][1])
    os.chdir(clips_path)
    clips.write_videofile(tick)


# In[17]:


# codes of following blocks used to convert mp4 files into mp3 files


# In[61]:


# create a list of mp3 file paths, used to store the converted mp3 files
mp3_filelist = []
audio_root = '/Users/xiteng/Desktop/speech-emotion-webapp-master/large audios'
for root, dirs, files in os.walk(video_path):  # 对文件夹进行遍历
    for name in files:
        if ".xls" not in name:
            if ".DS_Store" not in name:
                name = name.replace('.MP4', '.mp3')
                audio_path = audio_root+'/'+name
                mp3_filelist.append(audio_path)
for f in mp3_filelist:
    print(f)

#create a list of mp4 file paths, where the MP4 files are stored
mp4_filelist = []
for root, dirs, files in os.walk(video_path):  # 对文件夹进行遍历
    for name in files:
        if ".xls" not in name:
            if ".DS_Store" not in name:
                mp4_filelist.append(os.path.join(root, name))
for g in mp4_filelist:
    print(g)


# In[62]:


# converted mp4 to mp3, using path lists created above as inputs
for mp4_file, mp3_file in zip(mp4_filelist, mp3_filelist):
    videoclip = VideoFileClip(mp4_file)

    audioclip = videoclip.audio
    audioclip.write_audiofile(mp3_file)

    audioclip.close()
    videoclip.close()


# In[ ]:


#codes of following blocks used to get clips from audio files


# In[63]:


# create audio file list
audio_lst = []
for i in lst:
    audio_lst.append(i.replace('MP4', 'mp3'))
print(audio_lst)

#create a dictionary with keys=names, values=start time, end time
#audio version
audio_dict = {}
for tick, start, end in zip(audio_lst, time_starts, time_ends):
    audio_dict[tick] = [start, end]
print(audio_dict)


# In[65]:


#transform downloaded audios into audio clips 
audio_path = '/Users/xiteng/Desktop/speech-emotion-webapp-master/large audios'
audio_clips_path = '/Users/xiteng/Desktop/speech-emotion-webapp-master/clips'
for tick in audio_dict:
    os.chdir(audio_path)
    clips = AudioFileClip(tick).subclip(audio_dict[tick][0],audio_dict[tick][1])
    os.chdir(audio_clips_path)
    clips.write_audiofile(tick)

